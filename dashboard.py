"""
Tablero Ejecutivo de Operaciones - Centro de Distribucion (Consumo Masivo)
---------------------------------------------------------------------------
Lee el archivo Tablero_Ejecutivo_CD.xlsx y lo despliega como un dashboard
interactivo estilo "sala de control", igual al mockup de referencia.

Como correrlo:
1) pip install -r requirements.txt
2) streamlit run dashboard.py

El Excel se puede editar libremente (celdas amarillas: Valor Hoy, Meta,
Acum. Mes) y basta con recargar el navegador (o tocar el boton
"Actualizar datos") para ver los cambios reflejados.
"""

import io
import math
import random
from datetime import datetime

import openpyxl
import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st

# =========================================================================
# CONFIG GENERAL
# =========================================================================
st.set_page_config(
    page_title="Tablero Ejecutivo CD - Consumo Masivo",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ID de tu Google Sheet (lo que va entre /d/ y /edit en la URL).
# La hoja debe estar compartida como "Cualquiera con el enlace: Lector".
GOOGLE_SHEET_ID = "1JrdC8TNpJZB1qmp1sxY-g2iE9RDnun-BKmxlTrcydWU"
GOOGLE_SHEET_EXPORT_URL = (
    f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=xlsx"
)

COLOR_BG = "#0b1622"
COLOR_CARD = "#101d2c"
COLOR_CARD_BORDER = "#1c2f42"
COLOR_TEXT = "#e6edf3"
COLOR_MUTED = "#8fa3b8"
COLOR_GREEN = "#2ecc71"
COLOR_YELLOW = "#f1c40f"
COLOR_RED = "#e74c3c"
COLOR_ACCENT = "#3aa0ff"

ESTADO_COLOR = {
    "Cumple": COLOR_GREEN,
    "Cerca": COLOR_YELLOW,
    "No cumple": COLOR_RED,
}
ESTADO_ICON = {"Cumple": "✅", "Cerca": "⚠️", "No cumple": "❌"}


# =========================================================================
# HELPERS NUMERICOS (blindaje contra celdas vacias / NaN)
# =========================================================================
def is_number(val):
    """True solo si val es un numero real y finito (no None, no NaN, no inf)."""
    if val is None:
        return False
    if isinstance(val, bool):
        return False
    if isinstance(val, (int, float)):
        return not (math.isnan(val) or math.isinf(val))
    return False


def to_number(val, default=0):
    """Convierte a numero seguro; si no se puede (vacio, texto, NaN) devuelve default."""
    if is_number(val):
        return val
    try:
        f = float(val)
        return f if not (math.isnan(f) or math.isinf(f)) else default
    except (TypeError, ValueError):
        return default


# =========================================================================
# CSS - tema oscuro tipo "centro de control"
# =========================================================================
st.markdown(
    f"""
    <style>
    .stApp {{
        background-color: {COLOR_BG};
        color: {COLOR_TEXT};
    }}
    section[data-testid="stSidebar"] {{ display:none; }}
    #MainMenu, footer, header {{ visibility:hidden; }}
    .block-container {{ padding-top: 1.2rem; padding-bottom: 2rem; max-width: 1500px; }}
    .kpi-card {{
        background: {COLOR_CARD};
        border: 1px solid {COLOR_CARD_BORDER};
        border-radius: 10px;
        padding: 12px 14px;
        height: 130px;
    }}
    .kpi-title {{
        color: {COLOR_MUTED};
        font-size: 12px;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: .03em;
        margin-bottom: 4px;
    }}
    .kpi-value {{ font-size: 26px; font-weight: 700; color: {COLOR_TEXT}; }}
    .kpi-sub {{ font-size: 12px; color: {COLOR_MUTED}; margin-top: 6px; }}
    .kpi-meta {{ font-size: 12px; margin-top: 2px; }}
    .panel {{
        background: {COLOR_CARD};
        border: 1px solid {COLOR_CARD_BORDER};
        border-radius: 10px;
        padding: 10px 14px 14px 14px;
        margin-bottom: 14px;
    }}
    .panel-title {{
        color: {COLOR_ACCENT};
        font-size: 13px;
        font-weight: 700;
        letter-spacing: .04em;
        text-transform: uppercase;
        margin-bottom: 8px;
        border-bottom: 1px solid {COLOR_CARD_BORDER};
        padding-bottom: 6px;
    }}
    .row-line {{
        display:flex; justify-content:space-between; align-items:center;
        font-size: 13px; padding: 5px 2px; border-bottom: 1px solid #16283a;
    }}
    .row-line:last-child {{ border-bottom: none; }}
    .badge {{
        display:inline-block; width:10px; height:10px; border-radius:50%;
        margin-right:6px;
    }}
    .stat-box {{ text-align:center; }}
    .stat-num {{ font-size: 22px; font-weight:700; color:{COLOR_TEXT}; }}
    .stat-label {{ font-size: 11px; color:{COLOR_MUTED}; text-transform:uppercase; }}
    .header-title {{ font-size: 26px; font-weight:800; margin-bottom:0px;}}
    .header-sub {{ font-size: 12px; color:{COLOR_MUTED}; }}
    .alert-item {{ font-size: 13px; padding: 6px 2px; border-bottom: 1px solid #16283a; }}
    .alert-num {{
        display:inline-block; background:{COLOR_RED}; color:white; border-radius:50%;
        width:18px; height:18px; text-align:center; font-size:11px; margin-right:8px;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================================
# CARGA DE DATOS
# =========================================================================
@st.cache_data(ttl=15)
def fetch_workbook_bytes(url):
    """Descarga el libro completo de Google Sheets (todas las pestañas) como .xlsx."""
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    return resp.content


@st.cache_data(ttl=15)
def load_workbook(_unused_cache_key):
    raw = fetch_workbook_bytes(GOOGLE_SHEET_EXPORT_URL)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    def sheet_rows(name, start_row):
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row[0] is None:
                continue
            rows.append(row)
        return rows

    def kpi_table(name, start_row=5):
        rows = sheet_rows(name, start_row)
        data = []
        for r in rows:
            indicador, direccion, hoy, meta, cumplimiento, estado = r[:6]
            acum = r[6] if len(r) > 6 else None
            if indicador is None or estado is None or indicador == "Indicador":
                continue
            data.append(
                dict(
                    indicador=indicador,
                    direccion=direccion,
                    hoy=hoy,
                    meta=meta,
                    cumplimiento=cumplimiento,
                    estado=estado,
                    acum=acum,
                )
            )
        return pd.DataFrame(data)

    kpis = kpi_table("KPIs Principales")
    servicio = kpi_table("Servicio Cliente")
    productividad = kpi_table("Productividad")
    inventarios = kpi_table("Inventarios")
    costos = kpi_table("Costos")

    def daily_stats(name, n_stats):
        ws = wb[name]
        stats = []
        for row in ws.iter_rows(min_row=5, max_row=4 + n_stats, values_only=True):
            if row[0] is not None and row[1] is not None:
                stats.append((row[0], row[1]))
        return stats

    recepcion_stats = daily_stats("Recepcion", 3)
    recepcion_tabla = kpi_table("Recepcion", start_row=9)

    picking_stats = daily_stats("Picking", 3)
    picking_tabla = kpi_table("Picking", start_row=9)

    transporte_stats = daily_stats("Transporte", 3)
    transporte_tabla = kpi_table("Transporte", start_row=9)

    ecommerce_stats = daily_stats("Ecommerce", 3)
    ecommerce_tabla = kpi_table("Ecommerce", start_row=9)

    ws = wb["Andenes"]
    andenes_vals = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None and row[1] is not None:
            andenes_vals[row[0]] = row[1]

    ws = wb["Alertas"]
    alertas = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is not None and row[1] is not None:
            alertas.append((row[0], row[1]))

    ws = wb["Resumen Ejecutivo"]
    resumen = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is not None:
            resumen.append((row[0], row[1], row[2]))

    # Historico diario (opcional): indicador -> lista de valores ordenada por fecha.
    # Si el archivo aun no tiene esta hoja (version anterior), queda vacio y el
    # dashboard usa la tendencia sintetica como respaldo.
    historico = {}
    if "Historico" in wb.sheetnames:
        ws = wb["Historico"]
        raw = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            fecha, indicador, valor = (row + (None, None, None))[:3]
            if fecha is None or indicador is None or valor is None:
                continue
            raw.append((fecha, indicador, valor))
        raw.sort(key=lambda r: r[0])
        for fecha, indicador, valor in raw:
            historico.setdefault(indicador, []).append(valor)

    return dict(
        kpis=kpis,
        servicio=servicio,
        productividad=productividad,
        inventarios=inventarios,
        costos=costos,
        recepcion_stats=recepcion_stats,
        recepcion_tabla=recepcion_tabla,
        picking_stats=picking_stats,
        picking_tabla=picking_tabla,
        transporte_stats=transporte_stats,
        transporte_tabla=transporte_tabla,
        ecommerce_stats=ecommerce_stats,
        ecommerce_tabla=ecommerce_tabla,
        andenes=andenes_vals,
        alertas=alertas,
        resumen=resumen,
        historico=historico,
    )


def fmt_value(val, indicador):
    """Formatea el valor 'Hoy' segun el tipo de indicador (%, $, min, numero).

    Blindado contra celdas vacias, texto y NaN/inf: en esos casos muestra '-'
    en vez de reventar (antes 'int(NaN)' tiraba ValueError y caia toda la app).
    """
    indicador = indicador if isinstance(indicador, str) else str(indicador)
    ind = indicador.lower()

    # Celda vacia o valor no numerico valido -> guion.
    if not is_number(val):
        # Si es texto no vacio lo mostramos tal cual; si es None/NaN, guion.
        if isinstance(val, str) and val.strip():
            return val
        return "-"

    if "($)" in indicador or "costo" in ind or "$" in indicador:
        return f"${val:,.2f}"
    if 0 <= val <= 1 and (
        "%" in indicador
        or "otif" in ind
        or "nivel" in ind
        or "exactitud" in ind
        or "utilizaci" in ind
        or "quiebres" in ind
        or "rotaci" not in ind
        and "diferencias" not in ind
        and val <= 1
        and any(k in ind for k in ["completos", "servicio", "tiempo entrega", "cumplidas", "recepciones", "productividad", "reprocesos", "cancelaciones", "devoluciones", "da\u00f1os"])
    ):
        return f"{val*100:.1f}%"
    # float(val).is_integer() no revienta con NaN/inf (a diferencia de int(val)),
    # y ya sabemos que val es finito por is_number().
    if float(val).is_integer():
        return f"{int(val):,}"
    return f"{val:,.1f}"


def synth_trend(seed_text, end_value, n=7):
    """Genera una tendencia sintetica (placeholder) que termina en el valor de hoy.
    Sirve solo para la mini-grafica de 'Tendencia (7d)' mientras no exista
    un historico diario real en el Excel."""
    rnd = random.Random(str(seed_text))
    if not is_number(end_value):
        try:
            end = float(end_value)
        except (TypeError, ValueError):
            return None
        if math.isnan(end) or math.isinf(end):
            return None
    else:
        end = float(end_value)
    pts = [end * rnd.uniform(0.94, 1.06) for _ in range(n - 1)]
    pts.append(end)
    return pts


def get_trend(historico, indicador, end_value, n=7):
    """Devuelve los ultimos n valores reales de la hoja 'Historico' para este
    indicador. Si no hay suficiente historico (hoja ausente o indicador
    nuevo), cae de vuelta a la tendencia sintetica de siempre."""
    serie = historico.get(indicador)
    if serie and len(serie) >= 2:
        # Filtramos posibles vacios/NaN dentro del historico.
        limpia = [v for v in serie if is_number(v)]
        if len(limpia) >= 2:
            return limpia[-n:]
    return synth_trend(indicador, end_value, n=n)


def hex_to_rgba(hex_color, alpha=0.13):
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i : i + 2], 16) for i in (0, 2, 4))
    return f"rgba({r},{g},{b},{alpha})"


def sparkline(values, color):
    fig = go.Figure(
        go.Scatter(
            y=values,
            mode="lines",
            line=dict(color=color, width=2),
            fill="tozeroy",
            fillcolor=hex_to_rgba(color),
        )
    )
    fig.update_layout(
        margin=dict(l=0, r=0, t=0, b=0),
        height=32,
        width=110,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(visible=False),
        yaxis=dict(visible=False),
        showlegend=False,
    )
    return fig


# =========================================================================
# COMPONENTES DE RENDER
# =========================================================================
def render_kpi_card(col, row):
    estado = row["estado"]
    color = ESTADO_COLOR.get(estado, COLOR_MUTED)
    icon = ESTADO_ICON.get(estado, "")
    val_fmt = fmt_value(row["hoy"], row["indicador"])
    meta_fmt = fmt_value(row["meta"], row["indicador"])
    acum = row["acum"]
    acum_disp = acum if (acum is not None and str(acum).strip() != "") else "-"
    with col:
        st.markdown(
            f"""
            <div class="kpi-card">
                <div class="kpi-title">{row['indicador']}</div>
                <div class="kpi-value" style="color:{color}">{val_fmt}</div>
                <div class="kpi-meta">Meta: {meta_fmt} {icon}</div>
                <div class="kpi-sub">Acum. Mes: {acum_disp}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_table_panel(title, df, historico=None, n=7):
    historico = historico or {}
    st.markdown(f'<div class="panel"><div class="panel-title">{title}</div>', unsafe_allow_html=True)
    header_cols = st.columns([2.6, 1, 1, 0.6, 1.3])
    for c, h in zip(header_cols, ["Indicador", "Hoy", "Meta", "", f"Tendencia ({n}d)"]):
        c.markdown(f"<div style='color:{COLOR_MUTED};font-size:11px;font-weight:700;'>{h}</div>", unsafe_allow_html=True)
    for _, r in df.iterrows():
        color = ESTADO_COLOR.get(r["estado"], COLOR_MUTED)
        icon = ESTADO_ICON.get(r["estado"], "")
        c1, c2, c3, c4, c5 = st.columns([2.6, 1, 1, 0.6, 1.3])
        c1.markdown(f"<div style='font-size:12.5px;padding-top:4px;'>{r['indicador']}</div>", unsafe_allow_html=True)
        c2.markdown(f"<div style='font-size:12.5px;padding-top:4px;'>{fmt_value(r['hoy'], r['indicador'])}</div>", unsafe_allow_html=True)
        c3.markdown(f"<div style='font-size:12.5px;padding-top:4px;color:{COLOR_MUTED}'>{fmt_value(r['meta'], r['indicador'])}</div>", unsafe_allow_html=True)
        c4.markdown(f"<div style='padding-top:4px;'>{icon}</div>", unsafe_allow_html=True)
        vals = get_trend(historico, r["indicador"], r["hoy"], n=n)
        if vals:
            c5.plotly_chart(sparkline(vals, color), use_container_width=False, config={"displayModeBar": False}, key=f"spark-{title}-{r['indicador']}")
    st.markdown("</div>", unsafe_allow_html=True)


def render_area_card(title, stats, df, extra_html=""):
    st.markdown(f'<div class="panel"><div class="panel-title">{title}</div>', unsafe_allow_html=True)
    if stats:
        cols = st.columns(len(stats))
        for c, (label, val) in zip(cols, stats):
            # Formato robusto: porcentaje si es fraccion, entero si aplica, guion si no es numero.
            if is_number(val):
                if 0 < val < 1.5 and float(val) != int(val):
                    disp = f"{val*100:.1f}%"
                elif float(val).is_integer():
                    disp = f"{int(val):,}"
                else:
                    disp = f"{val:,.1f}"
            else:
                disp = val if (isinstance(val, str) and val.strip()) else "-"
            c.markdown(
                f"""<div class="stat-box"><div class="stat-num">{disp}</div>
                <div class="stat-label">{label}</div></div>""",
                unsafe_allow_html=True,
            )
    if extra_html:
        st.markdown(extra_html, unsafe_allow_html=True)
    for _, r in df.iterrows():
        color = ESTADO_COLOR.get(r["estado"], COLOR_MUTED)
        icon = ESTADO_ICON.get(r["estado"], "")
        st.markdown(
            f"""<div class="row-line">
                <span>{r['indicador']}</span>
                <span>{fmt_value(r['hoy'], r['indicador'])} &nbsp; {icon}</span>
            </div>""",
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)


def render_andenes_panel(data):
    st.markdown('<div class="panel"><div class="panel-title">ANDENES / MUELLES</div>', unsafe_allow_html=True)
    andenes = data["andenes"]
    c1, c2 = st.columns(2)
    c1.markdown(f"<div class='stat-box'><div class='stat-num'>{int(to_number(andenes.get('Total Andenes'), 0))}</div><div class='stat-label'>Total</div></div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='stat-box'><div class='stat-num'>{int(to_number(andenes.get('En Uso'), 0))}</div><div class='stat-label'>En Uso</div></div>", unsafe_allow_html=True)
    st.plotly_chart(render_andenes_donut(andenes), use_container_width=True, config={"displayModeBar": False}, key="donut-andenes")
    st.markdown("</div>", unsafe_allow_html=True)


def render_andenes_donut(andenes):
    labels = ["En Uso", "En Espera", "Fuera de Servicio", "Disponibles"]
    colors = [COLOR_ACCENT, "#e67e22", COLOR_RED, "#34495e"]
    values = [to_number(andenes.get(l), 0) for l in labels]

    # % Utilizacion: puede venir como fraccion (0.78), como porcentaje (78) o como texto.
    util_raw = andenes.get("% Utilización (calculado)")
    if is_number(util_raw):
        util = f"{util_raw*100:.0f}%" if util_raw <= 1 else f"{util_raw:.0f}%"
    elif isinstance(util_raw, str) and util_raw.strip():
        util = util_raw
    else:
        util = "-"

    fig = go.Figure(
        go.Pie(
            labels=labels,
            values=values,
            hole=0.62,
            marker=dict(colors=colors),
            textinfo="none",
        )
    )
    fig.update_layout(
        margin=dict(l=0, r=0, t=0, b=0),
        height=170,
        showlegend=False,
        paper_bgcolor="rgba(0,0,0,0)",
        annotations=[dict(text=f"<b>{util}</b><br><span style='font-size:10px'>Utilización</span>", x=0.5, y=0.5, showarrow=False, font=dict(color=COLOR_TEXT, size=16))],
    )
    return fig


# =========================================================================
# CARGA
# =========================================================================
try:
    data = load_workbook("v1")
except Exception as e:
    st.error(
        "No se pudo descargar la hoja de Google Sheets. Verifica que el link de "
        "GOOGLE_SHEET_ID sea correcto y que la hoja esté compartida como "
        f"'Cualquiera con el enlace: Lector'.\n\nDetalle: {e}"
    )
    st.stop()

# =========================================================================
# HEADER
# =========================================================================
h1, h2, h3, h4 = st.columns([3.2, 1, 1, 1.2])
with h1:
    st.markdown(
        """<div class="header-title">🚚 CENTRO DE DISTRIBUCIÓN — CONSUMO MASIVO</div>
        <div class="header-sub">TABLERO EJECUTIVO DE OPERACIONES</div>""",
        unsafe_allow_html=True,
    )

CATEGORIAS_TABLA = {
    "Servicio al Cliente": ("SERVICIO AL CLIENTE", "servicio"),
    "Productividad Operativa": ("PRODUCTIVIDAD OPERATIVA", "productividad"),
    "Inventarios": ("INVENTARIOS", "inventarios"),
    "Costos": ("COSTOS", "costos"),
}
CATEGORIAS_AREA = ["Recepción", "Picking", "Andenes / Muelles", "Transporte y Distribución", "E-commerce"]

with h2:
    categoria = st.selectbox(
        "Categoría",
        ["Todos"] + list(CATEGORIAS_TABLA.keys()) + CATEGORIAS_AREA,
        index=0, label_visibility="collapsed",
    )
with h3:
    periodo_label = st.selectbox(
        "Periodo", ["Hoy", "Últimos 7 días", "Últimos 14 días", "Últimos 30 días", "Últimos 90 días"],
        index=1, label_visibility="collapsed",
    )
PERIODO_DIAS = {"Hoy": 1, "Últimos 7 días": 7, "Últimos 14 días": 14, "Últimos 30 días": 30, "Últimos 90 días": 90}
n_dias = PERIODO_DIAS[periodo_label]

with h4:
    st.markdown(f"<div class='header-sub' style='text-align:right;'>Actualización:<br><b>{datetime.now().strftime('%d/%m/%Y %H:%M')}</b></div>", unsafe_allow_html=True)
    if st.button("🔄 Actualizar datos", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

st.markdown("<br>", unsafe_allow_html=True)

# =========================================================================
# FILA DE KPIs PRINCIPALES
# =========================================================================
if len(data["kpis"]) > 0:
    kpi_cols = st.columns(len(data["kpis"]))
    for col, (_, row) in zip(kpi_cols, data["kpis"].iterrows()):
        render_kpi_card(col, row)
else:
    st.info("La hoja 'KPIs Principales' no tiene datos legibles todavía.")

st.markdown("<br>", unsafe_allow_html=True)

# =========================================================================
# 4 PANELES: SERVICIO / PRODUCTIVIDAD / INVENTARIOS / COSTOS
# =========================================================================
tablas_visibles = (
    list(CATEGORIAS_TABLA.items())
    if categoria == "Todos"
    else [(categoria, CATEGORIAS_TABLA[categoria])] if categoria in CATEGORIAS_TABLA else []
)

if tablas_visibles:
    cols = st.columns(len(tablas_visibles))
    for col, (_, (titulo, key)) in zip(cols, tablas_visibles):
        with col:
            render_table_panel(titulo, data[key], data["historico"], n=n_dias)

# =========================================================================
# 5 AREAS OPERATIVAS
# =========================================================================
areas_visibles = CATEGORIAS_AREA if categoria == "Todos" else ([categoria] if categoria in CATEGORIAS_AREA else [])

AREA_RENDER = {
    "Recepción": lambda: render_area_card("RECEPCIÓN", data["recepcion_stats"], data["recepcion_tabla"]),
    "Picking": lambda: render_area_card("PICKING", data["picking_stats"], data["picking_tabla"]),
    "Andenes / Muelles": lambda: render_andenes_panel(data),
    "Transporte y Distribución": lambda: render_area_card("TRANSPORTE Y DISTRIBUCIÓN", data["transporte_stats"], data["transporte_tabla"]),
    "E-commerce": lambda: render_area_card("E-COMMERCE", data["ecommerce_stats"], data["ecommerce_tabla"]),
}

if areas_visibles:
    cols = st.columns(len(areas_visibles))
    for col, nombre in zip(cols, areas_visibles):
        with col:
            AREA_RENDER[nombre]()

if not tablas_visibles and not areas_visibles:
    st.info(f"'{categoria}' no tiene un panel propio; usa 'Todos' para ver el tablero completo.")

# =========================================================================
# ALERTAS + RESUMEN EJECUTIVO + TENDENCIA
# =========================================================================
b1, b2, b3 = st.columns([1.1, 1.6, 1.3])

with b1:
    st.markdown('<div class="panel"><div class="panel-title">⚠️ Top 5 Alertas</div>', unsafe_allow_html=True)
    for num, texto in data["alertas"]:
        num_disp = int(to_number(num, 0)) if is_number(num) else num
        st.markdown(
            f"""<div class="alert-item"><span class="alert-num">{num_disp}</span>{texto}</div>""",
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)

with b2:
    st.markdown('<div class="panel"><div class="panel-title">Tendencia Principales KPI</div>', unsafe_allow_html=True)
    st.caption(f"{periodo_label}, tomados de la hoja 'Histórico' del Excel.")
    fig = go.Figure()
    trend_series = [
        ("OTIF (%)", data["kpis"].loc[data["kpis"]["indicador"].str.contains("OTIF", case=False), "indicador"].values,
         data["kpis"].loc[data["kpis"]["indicador"].str.contains("OTIF", case=False), "hoy"].values, COLOR_ACCENT),
        ("Costo Op. x Caja ($)", data["kpis"].loc[data["kpis"]["indicador"].str.contains("Costo Operativo", case=False), "indicador"].values,
         data["kpis"].loc[data["kpis"]["indicador"].str.contains("Costo Operativo", case=False), "hoy"].values, COLOR_RED),
    ]
    for name, ind_arr, arr, color in trend_series:
        if len(arr):
            ys = get_trend(data["historico"], ind_arr[0], arr[0], n=n_dias)
            if ys:
                fig.add_trace(go.Scatter(y=ys, mode="lines+markers", name=name, line=dict(color=color, width=2), marker=dict(size=3)))
    fig.update_layout(
        height=230,
        margin=dict(l=10, r=10, t=10, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=COLOR_TEXT, size=11),
        legend=dict(orientation="h", y=-0.2),
        xaxis=dict(showgrid=False),
        yaxis=dict(showgrid=True, gridcolor="#16283a"),
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False}, key="trend-main")
    st.markdown("</div>", unsafe_allow_html=True)

with b3:
    st.markdown('<div class="panel"><div class="panel-title">Resumen Ejecutivo</div>', unsafe_allow_html=True)
    for indicador, valor, estado in data["resumen"]:
        color = ESTADO_COLOR.get(estado, COLOR_MUTED)
        icon = ESTADO_ICON.get(estado, "")
        valor_disp = fmt_value(valor, indicador)
        st.markdown(
            f"""<div class="row-line">
                <span>{indicador}</span>
                <span style="color:{color};font-weight:700;">{valor_disp} {icon}</span>
            </div>""",
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)

st.caption("* Los datos se leen en vivo desde tu Google Sheet (se refrescan solos cada ~15s, o al presionar 'Actualizar datos'). Edita las celdas amarillas directo en Google Sheets.")
